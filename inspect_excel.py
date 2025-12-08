
import openpyxl
import sys

file_path = "/Users/stalin_j/Skillboard/backend/app/static/sample_templates/sample_template.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    for sheet in wb.sheetnames:
        print(f"\n--- Sheet: {sheet} ---")
        ws = wb[sheet]
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            print(row)
            row_count += 1
            if row_count > 20:
                break
except Exception as e:
    print(f"Error reading file: {e}")
