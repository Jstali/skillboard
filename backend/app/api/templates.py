from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from typing import List, Any
from pydantic import BaseModel
import json
import csv
import os
import io
from datetime import datetime
import openpyxl
from pyexcel_ods3 import get_data

from app.db.database import get_db
from app.db.models import SkillTemplate, User
from app.api.dependencies import get_current_user

router = APIRouter(prefix="/api/admin/templates", tags=["templates"])


@router.post("/upload")
async def upload_template_file(
    file: UploadFile = File(...),
    template_name: str = Form(None),  # Optional custom name
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Upload a spreadsheet file (.xlsx, .csv, .ods) and extract all sheets as templates.
    Validates structure (must contain 'Skill' or 'Name' column).
    """
    print(f"DEBUG: Upload request received for file: {file.filename}")
    
    if not current_user.is_admin:
        print("DEBUG: User is not admin")
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Validate file extension
    file_ext = file.filename.lower().split('.')[-1] if '.' in file.filename else ''
    print(f"DEBUG: File extension: {file_ext}")
    
    if file_ext not in ['xlsx', 'csv', 'ods']:
        raise HTTPException(
            status_code=400,
            detail="Invalid file format. Supported formats: .xlsx, .csv, .ods"
        )
    
    try:
        # Read file content
        print("DEBUG: Reading file content...")
        content = await file.read()
        print(f"DEBUG: File content read, size: {len(content)} bytes")
        
        # Helper to validate structure
        def validate_structure(rows, source_name):
            if not rows or len(rows) < 1:
                return False
            
            # Look for header row in first 10 rows
            found_header = False
            for i in range(min(10, len(rows))):
                row_values = [str(c).lower().strip() for c in rows[i] if c]
                # Check for key columns
                if any(k in row_values for k in ['skill', 'name', 'competency', 'technical skills']):
                    found_header = True
                    break
            
            if not found_header:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Template not supported: '{source_name}' does not contain a valid 'Skill' or 'Name' column header."
                )
            return True

        templates_created = []
        
        if file_ext == 'xlsx':
            # Parse Excel file
            print("DEBUG: Loading Excel workbook...")
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                print(f"DEBUG: Workbook loaded. Sheets: {workbook.sheetnames}")
            except Exception as e:
                print(f"DEBUG: Error loading workbook: {e}")
                raise
            
            for sheet_name in workbook.sheetnames:
                print(f"DEBUG: Processing sheet: {sheet_name}")
                sheet = workbook[sheet_name]
                
                # Extract all rows and columns
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    # Convert row tuple to list, handling None values
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    rows.append(row_data)
                
                print(f"DEBUG: Sheet {sheet_name} has {len(rows)} rows")
                
                # Skip empty sheets
                if not rows or all(not any(cell for cell in row) for row in rows):
                    print(f"DEBUG: Skipping empty sheet {sheet_name}")
                    continue
                
                # Validate Structure
                print(f"DEBUG: Validating structure for {sheet_name}")
                validate_structure(rows, sheet_name)
                
                # Determine Template Name - format by replacing underscores with spaces
                formatted_sheet_name = sheet_name.replace('_', ' ')
                final_name = formatted_sheet_name
                if template_name:
                   final_name = template_name if len(workbook.sheetnames) == 1 else f"{template_name} - {formatted_sheet_name}"

                # Create template record
                print(f"DEBUG: Creating template record: {final_name}")
                template = SkillTemplate(
                    template_name=final_name,
                    file_name=file.filename,
                    content=json.dumps(rows),
                    uploaded_by=current_user.id
                )
                db.add(template)
                templates_created.append(final_name)
        
        elif file_ext == 'csv':
            # Parse CSV file
            csv_content = content.decode('utf-8-sig')  # Handle BOM
            csv_reader = csv.reader(io.StringIO(csv_content))
            
            rows = []
            for row in csv_reader:
                rows.append(row)
            
            if rows:
                validate_structure(rows, file.filename)
                
                # CSV files have only one "sheet", use filename or custom name
                # Format by replacing underscores with spaces
                base_name = template_name if template_name else file.filename.rsplit('.', 1)[0]
                final_name = base_name.replace('_', ' ')
                
                template = SkillTemplate(
                    template_name=final_name,
                    file_name=file.filename,
                    content=json.dumps(rows),
                    uploaded_by=current_user.id
                )
                db.add(template)
                templates_created.append(final_name)
        
        elif file_ext == 'ods':
            # Parse ODS file
            ods_data = get_data(io.BytesIO(content))
            
            for sheet_name, sheet_data in ods_data.items():
                # Skip empty sheets
                if not sheet_data or all(not any(cell for cell in row) for row in sheet_data):
                    continue
                
                # Convert all cells to strings
                rows = []
                for row in sheet_data:
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    rows.append(row_data)
                
                validate_structure(rows, sheet_name)

                # Format template name: replace underscores with spaces
                formatted_sheet_name = sheet_name.replace('_', ' ')
                final_name = formatted_sheet_name
                if template_name:
                     final_name = template_name if len(ods_data) == 1 else f"{template_name} - {formatted_sheet_name}"

                template = SkillTemplate(
                    template_name=final_name,
                    file_name=file.filename,
                    content=json.dumps(rows),
                    uploaded_by=current_user.id
                )
                db.add(template)
                templates_created.append(final_name)
        
        db.commit()
        
        return {
            "message": f"Successfully uploaded {len(templates_created)} template(s)",
            "templates_created": len(templates_created),
            "template_names": templates_created
        }
    
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process file: {str(e)}"
        )


@router.get("/sample")
async def get_sample_template(
    download: bool = False,
    template_name: str = None,
    current_user: User = Depends(get_current_user)
):
    """
    Get the sample template.
    If download=True:
        - If template_name provided: returns .xlsx with ONLY that sheet.
        - Else: returns full sample .xlsx file.
    Otherwise, returns the parsed content (list of sheets/templates).
    """
    sample_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "app", "static", "sample_templates", "sample_template.xlsx")
    
    
    # Helper to identify reference rows
    # Helper to identify reference rows
    def is_reference_row(row_values):
        # Strict keywords (Meta-headers)
        strict_keywords = ["Definition:", "Indicators:", "Typical Activities:"]
        
        # Specific phrases from the Rubric/Reference block to filter out
        # Using longer phrases to avoid accidental deletion of valid skill descriptions
        rubric_phrases = [
            "Requires close supervision",
            "Understanding of basic concepts",
            "Beginning to apply", 
            "Can perform routine tasks",
            "Shows initiative", 
            "May mentor junior",
            "Coaches other",
            "Trusted to lead",
            "Shapes strategy",
            "Recognised authority",
            "Drives innovation",
            "Represents the organisation"
        ]
        
        for cell in row_values:
            if cell and isinstance(cell, str):
                c = cell.strip()
                # Check strict keywords
                for k in strict_keywords:
                    if c.startswith(k):
                        return True
                # Check rubric phrases (content match)
                for p in rubric_phrases:
                    # Check if the cell starts with or strongly matches the rubric phrase
                    if c.startswith(p):
                        return True
        return False

    if download:
        from fastapi.responses import FileResponse, StreamingResponse
        
        if template_name:
            # Create a new workbook with just the requested sheet
            try:
                wb_source = openpyxl.load_workbook(sample_path, data_only=True)
                if template_name not in wb_source.sheetnames:
                    raise HTTPException(status_code=404, detail=f"Template '{template_name}' not found in sample file")
                
                source_sheet = wb_source[template_name]
                
                wb_dest = openpyxl.Workbook()
                dest_sheet = wb_dest.active
                dest_sheet.title = template_name
                
                for row in source_sheet.iter_rows(values_only=True):
                    # Filter reference rows
                    if not is_reference_row(row):
                        dest_sheet.append(row)
                
                # Save to memory
                output = io.BytesIO()
                wb_dest.save(output)
                output.seek(0)
                
                filename = f"{template_name.replace(' ', '_')}.xlsx"
                return StreamingResponse(
                    output, 
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'},
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Failed to extract template: {str(e)}")

        return FileResponse(
            sample_path, 
            filename="Skillboard_Sample_Templates_All.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    # Parse and return content
    try:
        with open(sample_path, "rb") as f:
            content = f.read()
            
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        templates = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                # Clean row data
                row_data = [str(cell) if cell is not None else '' for cell in row]
                
                # Check for reference rows
                if is_reference_row(row_data):
                    continue
                    
                rows.append(row_data)
            
            if rows and any(any(row) for row in rows):
               templates.append({
                   "template_name": sheet_name,
                   "content": rows
               })
               
        return templates

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read sample: {str(e)}")

@router.get("")
async def get_all_templates(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all uploaded templates with metadata (without full content)."""
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    templates = db.query(SkillTemplate).order_by(SkillTemplate.created_at.desc()).all()
    
    return [
        {
            "id": t.id,
            "template_name": t.template_name,
            "file_name": t.file_name,
            "created_at": t.created_at.isoformat(),
            "row_count": len(json.loads(t.content)) if t.content else 0
        }
        for t in templates
    ]


@router.get("/{template_id}")
async def get_template_by_id(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get a specific template with full content for viewing."""
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    template = db.query(SkillTemplate).filter(SkillTemplate.id == template_id).first()
    
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    return {
        "id": template.id,
        "template_name": template.template_name,
        "file_name": template.file_name,
        "created_at": template.created_at.isoformat(),
        "content": json.loads(template.content)
    }


class TemplateUpdate(BaseModel):
    content: List[List[Any]]

@router.put("/{template_id}")
async def update_template(
    template_id: int,
    template_update: TemplateUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Update template content.
    Only admins can update templates.
    """
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")

    template = db.query(SkillTemplate).filter(SkillTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    template.content = json.dumps(template_update.content)
    
    db.commit()
    db.refresh(template)
    
    return {
        "id": template.id,
        "template_name": template.template_name,
        "file_name": template.file_name,
        "created_at": template.created_at.isoformat(),
        "content": json.loads(template.content)
    }


@router.delete("/{template_id}")
async def delete_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a specific template."""
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    template = db.query(SkillTemplate).filter(SkillTemplate.id == template_id).first()
    
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    template_name = template.template_name
    
    try:
        # Manually cascade delete assignments
        # Use simple delete query for performance
        from app.db.models import TemplateAssignment
        db.query(TemplateAssignment).filter(TemplateAssignment.template_id == template_id).delete()
        
        db.delete(template)
        db.commit()
    except Exception as e:
        db.rollback()
        # Log the error preferably
        print(f"Error deleting template: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Failed to delete template: {str(e)}")
    
    return {
        "message": f"Template '{template_name}' deleted successfully"
    }


class TemplateRename(BaseModel):
    new_name: str

@router.put("/{template_id}/rename")
async def rename_template(
    template_id: int,
    rename_data: TemplateRename,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Rename a template."""
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    template = db.query(SkillTemplate).filter(SkillTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    template.template_name = rename_data.new_name
    db.commit()
    db.refresh(template)
    
    return template


@router.get("/options/dropdown")
async def get_templates_dropdown(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get lightweight list of templates for dropdowns."""
    templates = db.query(SkillTemplate.id, SkillTemplate.template_name).order_by(SkillTemplate.template_name).all()
    return [{"id": t.id, "name": t.template_name} for t in templates]


class AssignmentRequest(BaseModel):
    template_id: int
    employee_ids: List[int] = []
    department: str = None
    role: str = None
    team: str = None

@router.post("/assign")
async def assign_template_to_employees(
    assignment: AssignmentRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Assign a template to employees.
    Can assign to specific employee IDs, OR bulk assign by department/role/team.
    """
    print(f"DEBUG: Assignment request received")
    print(f"DEBUG: template_id={assignment.template_id}")
    print(f"DEBUG: employee_ids={assignment.employee_ids}")
    print(f"DEBUG: department={assignment.department}")
    print(f"DEBUG: role={assignment.role}")
    print(f"DEBUG: team={assignment.team}")
    
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.db.models import TemplateAssignment, Employee
    
    # Verify template exists
    template = db.query(SkillTemplate).filter(SkillTemplate.id == assignment.template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Resolve target employees
    target_employees = []
    
    if assignment.employee_ids:
        # Direct selection
        print(f"DEBUG: Querying employees with IDs: {assignment.employee_ids}")
        target_employees = db.query(Employee).filter(Employee.id.in_(assignment.employee_ids)).all()
        print(f"DEBUG: Found {len(target_employees)} employees")
    
    elif assignment.department or assignment.role or assignment.team:
        # Filter based selection
        query = db.query(Employee)
        
        if assignment.department:
            query = query.filter(Employee.department == assignment.department)
        if assignment.role:
            query = query.filter(Employee.role == assignment.role)
        if assignment.team:
            query = query.filter(Employee.team == assignment.team)
            
        target_employees = query.all()
    
    if not target_employees:
        return {
            "message": "No employees found matching criteria",
            "assignments_created": 0,
            "errors": []
        }

    assignments_created = 0
    errors = []
    skipped = 0
    
    for employee in target_employees:
        # Check if already assigned
        existing = db.query(TemplateAssignment).filter(
            TemplateAssignment.template_id == assignment.template_id,
            TemplateAssignment.employee_id == employee.id
        ).first()
        
        if existing:
            skipped += 1
            errors.append(f"{employee.name} already has this template assigned")
            continue
            
        new_assignment = TemplateAssignment(
            template_id=assignment.template_id,
            employee_id=employee.id,
            assigned_by=current_user.id
        )
        db.add(new_assignment)
        assignments_created += 1
    
    db.commit()
    
    message = f"Assigned template to {assignments_created} employees"
    if skipped > 0:
        message += f" ({skipped} already assigned)"
    
    return {
        "message": message,
        "assignments_created": assignments_created,
        "skipped": skipped,
        "errors": errors
    }
