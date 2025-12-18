"""Script to import skills with band requirements from Excel.

Expected Excel format:
| Skill | A | B | C | L1 | L2 |
| 1. Category Name | Beginner | Developing | Intermediate | Advanced | Expert |
| Skill Name 1 | Beginner | Developing | Intermediate | Advanced | Expert |
| Skill Name 2 | Beginner | Developing | Intermediate | Advanced | Expert |
| 2. Another Category | ... | ... | ... | ... | ... |

Run with: docker exec skillboard-backend python import_skills_with_requirements.py <excel_file>
"""
import os
import sys
import re

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.db.database import SessionLocal
from app.db.models import Skill, RoleRequirement, RatingEnum


def parse_category_header(text: str) -> str | None:
    """Check if row is a category header like '1. Analytical & Problem Solving'"""
    if not text:
        return None
    # Match pattern: number followed by dot and category name
    match = re.match(r'^\d+\.\s*(.+)$', text.strip())
    if match:
        return match.group(1).strip()
    return None


def normalize_rating(rating: str) -> str | None:
    """Normalize rating string to enum value."""
    if not rating:
        return None
    rating = rating.strip().lower()
    mapping = {
        'beginner': 'Beginner',
        'developing': 'Developing',
        'intermediate': 'Intermediate',
        'advanced': 'Advanced',
        'expert': 'Expert',
    }
    return mapping.get(rating)


def import_from_excel(excel_path: str, delete_existing: bool = False):
    """Import skills and requirements from Excel file."""
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
    
    db = SessionLocal()
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        print(f"Headers: {headers}")
        
        # Find column indices
        skill_col = 0  # First column is skill name
        band_cols = {}
        for i, h in enumerate(headers):
            if h and h.upper() in ['A', 'B', 'C', 'L1', 'L2']:
                band_cols[h.upper()] = i
        
        print(f"Band columns: {band_cols}")
        
        if delete_existing:
            print("\nDeleting existing skills and requirements...")
            from app.db.models import EmployeeSkill, TeamSkillTemplate, CategorySkillTemplate, SkillGapResult, EmployeeTemplateResponse
            from sqlalchemy import text
            
            db.query(EmployeeTemplateResponse).delete()
            db.query(SkillGapResult).delete()
            db.query(EmployeeSkill).delete()
            db.query(TeamSkillTemplate).delete()
            db.query(RoleRequirement).delete()
            db.query(CategorySkillTemplate).delete()
            db.execute(text("UPDATE courses SET skill_id = NULL"))
            db.execute(text("UPDATE course_assignments SET skill_id = NULL"))
            db.query(Skill).delete()
            db.commit()
            print("✓ Deleted existing data")
        
        # Parse rows
        current_category = None
        skills_created = 0
        requirements_created = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            skill_name = row[skill_col]
            if not skill_name:
                continue
            
            skill_name = str(skill_name).strip()
            
            # Check if this is a category header
            category = parse_category_header(skill_name)
            if category:
                current_category = category
                print(f"\n=== Category: {current_category} ===")
                continue
            
            # Skip if no category yet
            if not current_category:
                continue
            
            # This is a skill row
            print(f"  Skill: {skill_name}")
            
            # Create or get skill
            existing_skill = db.query(Skill).filter(Skill.name == skill_name).first()
            if existing_skill:
                skill = existing_skill
                skill.category = current_category
            else:
                skill = Skill(
                    name=skill_name,
                    category=current_category,
                    description=f"{current_category} skill"
                )
                db.add(skill)
                db.flush()  # Get the ID
                skills_created += 1
            
            # Create role requirements for each band
            for band, col_idx in band_cols.items():
                rating_str = row[col_idx] if col_idx < len(row) else None
                rating = normalize_rating(str(rating_str) if rating_str else '')
                
                if rating:
                    # Check if requirement exists
                    existing_req = db.query(RoleRequirement).filter(
                        RoleRequirement.band == band,
                        RoleRequirement.skill_id == skill.id
                    ).first()
                    
                    if existing_req:
                        existing_req.required_rating = RatingEnum(rating)
                    else:
                        req = RoleRequirement(
                            band=band,
                            skill_id=skill.id,
                            required_rating=RatingEnum(rating),
                            is_required=True
                        )
                        db.add(req)
                        requirements_created += 1
        
        db.commit()
        
        print(f"\n✓ Created {skills_created} skills")
        print(f"✓ Created {requirements_created} role requirements")
        
        # Show summary
        print("\n=== Skills by Category ===")
        from collections import defaultdict
        skills = db.query(Skill).order_by(Skill.category, Skill.name).all()
        by_cat = defaultdict(list)
        for s in skills:
            by_cat[s.category or 'No Category'].append(s.name)
        
        for cat, lst in sorted(by_cat.items()):
            print(f"  {cat}: {len(lst)} skills")
        
        print(f"\nTotal skills: {len(skills)}")
        print(f"Total requirements: {db.query(RoleRequirement).count()}")
        
    except Exception as e:
        db.rollback()
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_skills_with_requirements.py <excel_file> [--delete-existing]")
        print("Example: python import_skills_with_requirements.py /app/skills.xlsx --delete-existing")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    delete_existing = "--delete-existing" in sys.argv
    
    if delete_existing:
        confirm = input("This will DELETE all existing skills. Type 'YES' to confirm: ")
        if confirm != "YES":
            print("Aborted.")
            sys.exit(0)
    
    import_from_excel(excel_path, delete_existing)
